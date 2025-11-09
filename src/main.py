from flask import Flask, render_template, request, redirect, url_for, jsonify
from flask_pymongo import PyMongo
from forms import InventoryItemForm, TransactionForm
import datetime
from zoneinfo import ZoneInfo
from bson import ObjectId
import re
import io
from openpyxl import load_workbook
from flask import send_file

app = Flask(__name__)
app.config.from_object('config.Config')

# inicializamos mongo (si Flask-PyMongo está disponible)
mongo = None
try:
    mongo = PyMongo(app)
except Exception:
    # si falla la inicialización de Flask-PyMongo, dejamos mongo en None
    mongo = None

# intenta inicializar models.py si exporta init_db/check_db_connection/get_db
try:
    from models import find_transactions
    from models import init_db, check_db_connection, get_db as models_get_db, add_inventory_item, add_transaction, find_inventory
    try:
        init_db(app)
    except Exception:
        # init_db puede lanzar si ya fue inicializada o por configuración ausente
        pass
except ImportError:
    init_db = None
    check_db_connection = None
    models_get_db = None
    add_inventory_item = None
    add_transaction = None
    find_inventory = None

def _next_transaction_id():
    """Calcula el siguiente id de transacción basándose en la última transacción.
    Estrategia:
      - Busca el último documento en la colección transactions (orden por _id desc).
      - Extrae el número al final del campo 'id' si existe (ej. 'tx_12' -> 12) y lo incrementa.
      - Si no hay id numérico, usa el prefijo 'tx_' y empieza en 1.
    Nota: esto no garantiza atomicidad en concurrencia; para alta concurrencia usar una colección de contadores.
    """
    try:
        db = get_database()
    except Exception:
        return "1"

    last = db.transactions.find_one({}, sort=[('_id', -1)])
    if not last:
        return "1"
    last_id = last.get('id') or ""
    m = re.search(r'(\d+)$', str(last_id))
    if m:
        try:
            n = int(m.group(1)) + 1
            # mantener el prefijo tal cual (todo lo anterior al número)
            prefix = str(last_id)[:m.start(1)]
            return f"{prefix}{n}"
        except Exception:
            return f"{str(last_id)}_1"
    # si no tiene número al final, añadir sufijo _1
    return f"{str(last_id)}_1"

def _next_inventory_id():
    """Calcula el siguiente id de inventario basándose en el último documento.
    Estrategia:
      - Busca el último documento en la colección transactions (orden por _id desc).
      - Extrae el número al final del campo 'id' si existe (ej. 'tx_12' -> 12) y lo incrementa.
      - Si no hay id numérico, usa el prefijo 'tx_' y empieza en 1.
    Nota: esto no garantiza atomicidad en concurrencia; para alta concurrencia usar una colección de contadores.
    """
    try:
        db = get_database()
    except Exception:
        return "1"

    last = db.inventory.find_one({}, sort=[('_id', -1)])
    if not last:
        return "1"
    last_id = last.get('id') or ""
    m = re.search(r'(\d+)$', str(last_id))
    if m:
        try:
            n = int(m.group(1)) + 1
            # mantener el prefijo tal cual (todo lo anterior al número)
            prefix = str(last_id)[:m.start(1)]
            return f"{prefix}{n}"
        except Exception:
            return f"{str(last_id)}_1"
    # si no tiene número al final, añadir sufijo _1
    return f"{str(last_id)}_1"

def _serialize_docs(docs):
    """Convierte ObjectId y datetime a valores JSON/str para mostrar en templates."""
    out = []
    for d in docs:
        doc = {}
        for k, v in d.items():
            if isinstance(v, object):
                doc[k] = str(v)
            elif isinstance(v, (datetime.datetime, datetime.date)) and not isinstance(v, str):
                # convertir date->datetime->ISO
                if isinstance(v, datetime.date) and not isinstance(v, datetime.datetime):
                    v = datetime.datetime.combine(v, datetime.time.min)
                doc[k] = v.isoformat()
            else:
                doc[k] = v
        out.append(doc)
    return out

def get_database():
    """Devuelve un objeto Database usable. Intenta models.get_db() y luego flask-pymongo."""
    # primero usar get_db desde models.py si existe
    if 'models_get_db' in globals() and callable(models_get_db):
        try:
            db = models_get_db()
            if db is not None:
                return db
        except Exception:
            # si models.get_db falla o devuelve None, seguir al fallback
            pass
    # fallback a flask-pymongo
    if mongo is not None:
        try:
            db = getattr(mongo, 'db', None)
            if db is not None:
                return db
        except Exception:
            pass
    raise RuntimeError("No hay cliente Mongo configurado o la función get_db devolvió None")

def verify_mongo_connection():
    """Intenta verificar la conexión a MongoDB.
    Retorna (True, None) si ok, (False, error) si falla."""
    # preferir el método de models.py si existe
    if 'check_db_connection' in globals() and callable(check_db_connection):
        try:
            ok = check_db_connection()
            return True, None if ok else (False, "ping falló")
        except Exception as e:
            return False, e

    # fallback: intentar usar el cliente de flask_pymongo si está disponible
    if mongo is not None:
        try:
            # obtener objeto client/cliente real de PyMongo
            client = getattr(mongo, 'cx', None) or getattr(mongo, 'client', None) or getattr(mongo, 'mongo_client', None)
            # si tenemos una Database (mongo.db), obtener su .client
            if client is None:
                db_obj = getattr(mongo, 'db', None)
                client = getattr(db_obj, 'client', None) if db_obj is not None else None
            if client is None:
                return False, "No se pudo localizar cliente Mongo en flask_pymongo"
            pong = client.admin.command('ping')
            return (True, None) if pong.get('ok') == 1.0 else (False, "ping respuesta inesperada")
        except Exception as e:
            return False, e

    return False, "No hay cliente Mongo configurado"

@app.route('/db_check')
def db_check():
    ok, err = verify_mongo_connection()
    if ok:
        return "MongoDB OK", 200
    return f"MongoDB ERROR: {err}", 500

@app.route('/')
def index():
    # Página principal simple con enlaces a las vistas separadas
    return render_template('index.html')

@app.route('/inventory')
def inventory_view():
    inventory = []
    if 'find_inventory' in globals() and callable(find_inventory):
        try:
            inventory = find_inventory()
        except Exception as e:
            print("find_inventory error:", e)

    if not inventory:
        try:
            db = get_database()
            inventory = list(db.inventory.find().limit(200))
        except Exception as e:
            print("fallback inventory error:", e)
            inventory = []

    inventory = _serialize_docs(inventory)
    return render_template('inventory.html', inventory_list=inventory)

@app.route('/transactions')
def transactions_view():
    transactions = []
    if 'find_transactions' in globals() and callable(find_transactions):
        try:
            transactions = find_transactions(aggregate_by_product=False, group_by_code=False)
        except Exception as e:
            print("find_transactions error:", e)

    if not transactions:
        try:
            db = get_database()
            transactions = list(db.transactions.find().limit(200))
        except Exception as e:
            print("fallback transactions error:", e)
            transactions = []

    transactions = _serialize_docs(transactions)
    return render_template('transactions.html', transactions_list=transactions)

@app.route('/add_inventory', methods=['GET', 'POST'])
def add_inventory():
    form = InventoryItemForm()

    # sugerir id basado en la última entrada y mostrarlo en el formulario
    try:
        suggested_id = _next_inventory_id()
        form.id.data = suggested_id
    except Exception:
        suggested_id = None

    if request.method == 'POST':
        # recalcular justo antes de insertar para reducir colisiones
        try:
            form.id.data = _next_inventory_id()
        except Exception:
            pass

        item = {
            'id': form.id.data,
            'code': form.code.data,
            'product': form.product.data,
            'shelves': form.shelves.data,
            'floors': form.floors.data,
            'packs': form.packs.data
        }
        print("Adding inventory item:", item)
        try:
            inserted_id = add_inventory_item(item)
        except Exception as e:
            return f"MongoDB ERROR inserting inventory: {e}", 500

        print("Inserted inventory _id:", inserted_id)
        return redirect(url_for('inventory_view'))

    return render_template('add_inventory.html', form=form, suggested_id=suggested_id)

@app.route('/record_transaction', methods=['GET', 'POST'])
def record_transaction():
    form = TransactionForm()

    # obtener la fecha/hora actual en Colombia (zona America/Bogota)
    colombia_now = datetime.datetime.now(ZoneInfo("America/Bogota"))
    suggested_id = _next_transaction_id()
    
    try:
        form.id.data = suggested_id
    except Exception:
        pass

    try:
        form.date.data = colombia_now.date()
    except Exception:
        form.date.data = colombia_now

    # cargar opciones de inventario (incluye ahora 'code' y tratamos packs como decimal)
    inventory_options = []
    try:
        if 'find_inventory' in globals() and callable(find_inventory):
            inv_docs = find_inventory()
        else:
            db = get_database()
            inv_docs = list(db.inventory.find({}, {'product': 1, 'code': 1, 'shelves': 1, 'floors': 1, 'packs': 1}).limit(500))
    except Exception as e:
        print("Error cargando inventario para select:", e)
        inv_docs = []

    for doc in inv_docs:
        _id = str(doc.get('_id'))
        prod = doc.get('product') or doc.get('code') or _id
        inventory_options.append({
            'value': _id,
            'label': f"{prod}" + (f" — {doc.get('code')}" if doc.get('code') else ""),
            'shelves': doc.get('shelves', 1),
            'floors': doc.get('floors', 1),
            # packs pueden ser decimales (cajas)
            'packs': float(doc.get('packs', 1)),
            'code': doc.get('code', ''),
            'product_name': prod
        })

    if form.validate_on_submit():
        selected = form.product.data
        inv_doc = None
        if selected:
            try:
                oid = ObjectId(selected)
                if 'find_inventory' in globals() and callable(find_inventory):
                    found = find_inventory({'_id': oid})
                    inv_doc = found[0] if found else None
                else:
                    db = get_database()
                    inv_doc = db.inventory.find_one({'_id': oid})
            except Exception:
                inv_doc = None

        def to_num(v):
            try:
                return float(v)
            except Exception:
                return 0.0
                                    
        # campos temporales:
        # - temp_shelves / temp_floors: enteros multiplicados por sus multipliers en inventory
        # - temp_packs: multiplicado por inv_packs (packs en inventory)
        # - temp_cajas: DECIMAL que se suma directamente (no multiplicador)
        temp_shelves = to_num(request.form.get('temp_shelves', 0))
        temp_floors = to_num(request.form.get('temp_floors', 0))
        temp_packs = to_num(request.form.get('temp_packs', 0))
        temp_cajas = to_num(request.form.get('temp_cajas', 0))  # nuevas cajas decimales que se suman

        # multiplicadores desde inventario (aseguramos float)
        inv_shelves = float(inv_doc.get('shelves', 1)) if inv_doc else 1.0
        inv_floors = float(inv_doc.get('floors', 1)) if inv_doc else 1.0
        inv_packs = float(inv_doc.get('packs', 1)) if inv_doc else 1.0

        # calcular total:
        # total = shelves*inv_shelves + floors*inv_floors + packs*inv_packs + cajas (sin multiplicador)
        total_calc = (temp_shelves * inv_shelves) + (temp_floors * inv_floors) + (temp_packs ) + (temp_cajas/ inv_packs)

        # determinar nombre de producto final
        product_name = inv_doc.get('product') if inv_doc else form.product.data

        # determinar código: si viene del inventario, usarlo; si viene código manual, usarlo
        codigo = ''
        if inv_doc:
            codigo = inv_doc.get('code', '') or ''
        else:
            codigo = request.form.get('codigo_custom') or request.form.get('codigo') or ''

        # obtener referencia al _id del inventario (si se seleccionó)
        inventory_ref_id = inv_doc.get('_id') if inv_doc else None

        # construir transacción: usar colombia_now como fecha definitiva (no usar valor enviado por cliente)
        tx = {
            'id': form.id.data,
            'date': colombia_now,   # timezone-aware fecha/hora de Colombia
            'product': product_name,
            'codigo': codigo,
            'total': total_calc,
            # referencia al documento de inventory (ObjectId) — se añade solo si proviene del select
            'inventory_id': inventory_ref_id
        }

        try:
            if add_transaction and callable(add_transaction):
                inserted_id = add_transaction(tx)
            else:
                db = get_database()
                inserted_id = db.transactions.insert_one(tx).inserted_id
        except Exception as e:
            return f"MongoDB ERROR inserting transaction: {e}", 500

        print("Inserted transaction (computed):", tx)
        print("Inserted transaction _id:", inserted_id)
        return redirect(url_for('transactions_view'))

    # render con inventory_options y la fecha de Colombia para mostrar en el template
    return render_template('record_transaction.html', form=form, inventory_options=inventory_options, colombia_now=colombia_now)

@app.route('/delete_inventory/<id>', methods=['POST'])
def delete_inventory(id):
    """Eliminar por _id (ObjectId) o por campo 'id' si falla."""
    try:
        db = get_database()
    except Exception as e:
        return f"MongoDB ERROR: {e}", 500

    deleted_count = 0
    try:
        # intentar como ObjectId
        oid = ObjectId(id)
        res = db.inventory.delete_one({'_id': oid})
        deleted_count = res.deleted_count
    except Exception:
        # fallback a campo 'id' (string)
        try:
            res = db.inventory.delete_one({'id': id})
            deleted_count = res.deleted_count
        except Exception as e:
            return f"MongoDB ERROR deleting inventory: {e}", 500

    print(f"Deleted inventory id={id} deleted_count={deleted_count}")
    return redirect(url_for('inventory_view'))

@app.route('/delete_transaction/<id>', methods=['POST'])
def delete_transaction(id):
    """Eliminar transacción por _id (ObjectId) o por campo 'id'."""
    try:
        db = get_database()
    except Exception as e:
        return f"MongoDB ERROR: {e}", 500

    deleted_count = 0
    try:
        oid = ObjectId(id)
        res = db.transactions.delete_one({'_id': oid})
        deleted_count = res.deleted_count
    except Exception:
        try:
            res = db.transactions.delete_one({'id': id})
            deleted_count = res.deleted_count
        except Exception as e:
            return f"MongoDB ERROR deleting transaction: {e}", 500

    print(f"Deleted transaction id={id} deleted_count={deleted_count}")
    return redirect(url_for('transactions_view'))

@app.route('/edit_inventory/<id>', methods=['GET', 'POST'])
def edit_inventory(id):
    form = InventoryItemForm()
    # localizar documento por ObjectId o por campo 'id'
    db = None
    try:
        db = get_database()
    except Exception as e:
        return f"MongoDB ERROR: {e}", 500

    doc = None
    try:
        oid = ObjectId(id)
        doc = db.inventory.find_one({'_id': oid})
    except Exception:
        doc = db.inventory.find_one({'id': id})

    if not doc:
        return "Inventario no encontrado", 404

    if request.method == 'GET':
        # popular formulario con valores existentes
        try:
            form.id.data = doc.get('id')
            form.code.data = doc.get('code')
            form.product.data = doc.get('product')
            form.shelves.data = doc.get('shelves')
            form.floors.data = doc.get('floors')
            form.packs.data = doc.get('packs')
        except Exception:
            pass
        return render_template('edit_inventory.html', form=form, doc_id=str(doc.get('_id')))

    # POST: actualizar
    if form.validate_on_submit():
        update = {
            'id': form.id.data,
            'code': form.code.data,
            'product': form.product.data,
            'shelves': form.shelves.data,
            'floors': form.floors.data,
            'packs': form.packs.data
        }
        try:
            # intentar actualizar por ObjectId primero, si fracas, por campo id
            try:
                oid = ObjectId(id)
                res = db.inventory.update_one({'_id': oid}, {'$set': update})
            except Exception:
                res = db.inventory.update_one({'id': id}, {'$set': update})
            print("Inventory updated:", update, "modified_count:", getattr(res, 'modified_count', None))
        except Exception as e:
            return f"MongoDB ERROR updating inventory: {e}", 500
        return redirect(url_for('inventory_view'))
    return render_template('edit_inventory.html', form=form, doc_id=str(doc.get('_id')))

@app.route('/edit_transaction/<id>', methods=['GET', 'POST'])
def edit_transaction(id):
    form = TransactionForm()
    try:
        db = get_database()
    except Exception as e:
        return f"MongoDB ERROR: {e}", 500

    doc = None
    try:
        oid = ObjectId(id)
        doc = db.transactions.find_one({'_id': oid})
    except Exception:
        doc = db.transactions.find_one({'id': id})

    if not doc:
        return "Transacción no encontrada", 404

    if request.method == 'GET':
        # popular formulario (fecha se mostrará readonly en template)
        try:
            form.id.data = doc.get('id')
            # si el campo date es datetime, asignar date()
            dt = doc.get('date')
            try:
                form.date.data = dt.date() if hasattr(dt, 'date') else dt
            except Exception:
                form.date.data = dt
            form.product.data = doc.get('product')
            form.total.data = doc.get('total')
        except Exception:
            pass
        return render_template('edit_transaction.html', form=form, doc_id=str(doc.get('_id')), stored_date=doc.get('date'))

    # POST: actualizar transacción (no permitimos cambiar la fecha)
    if form.validate_on_submit():
        # intentar recalcular total si se enviaron temp_* campos, si no usar form.total
        def to_num(v):
            try:
                return float(v)
            except Exception:
                return 0.0

        temp_shelves = to_num(request.form.get('temp_shelves', 0))
        temp_floors = to_num(request.form.get('temp_floors', 0))
        temp_packs = to_num(request.form.get('temp_packs', 0))

        total_calc = None
        if (temp_shelves or temp_floors or temp_packs):
            # si el producto value corresponde a inventory _id, buscar multiplicadores
            inv_doc = None
            selected = form.product.data
            try:
                oid_sel = ObjectId(selected)
                inv_doc = db.inventory.find_one({'_id': oid_sel})
            except Exception:
                inv_doc = db.inventory.find_one({'product': selected}) or db.inventory.find_one({'id': selected})
            inv_shelves = inv_doc.get('shelves', 1) if inv_doc else 1
            inv_floors = inv_doc.get('floors', 1) if inv_doc else 1
            inv_packs = inv_doc.get('packs', 1) if inv_doc else 1
            total_calc = (temp_shelves * inv_shelves) + (temp_floors * inv_floors) + (temp_packs * inv_packs)

        new_total = total_calc if total_calc is not None else form.total.data

        # si form.product.data es ObjectId string, resolver nombre
        product_name = form.product.data
        try:
            oid_prod = ObjectId(form.product.data)
            inv_doc = db.inventory.find_one({'_id': oid_prod})
            if inv_doc:
                product_name = inv_doc.get('product') or product_name
        except Exception:
            pass

        update = {
            'id': form.id.data,
            # no tocamos 'date' para mantener fecha original
            'product': product_name,
            'total': new_total
        }

        try:
            try:
                oid = ObjectId(id)
                res = db.transactions.update_one({'_id': oid}, {'$set': update})
            except Exception:
                res = db.transactions.update_one({'id': id}, {'$set': update})
            print("Transaction updated:", update, "modified_count:", getattr(res, 'modified_count', None))
        except Exception as e:
            return f"MongoDB ERROR updating transaction: {e}", 500
        return redirect(url_for('transactions_view'))

    return render_template('edit_transaction.html', form=form, doc_id=str(doc.get('_id')), stored_date=doc.get('date'))

@app.route('/download_report')
def download_report():
    """
    Carga una plantilla Excel (xlsx) desde la carpeta 'excel_templates' y la rellena con:
      - Columna código ya existente en la hoja 'Datos'
      - Rellena la columna 'total' según los datos traídos desde la BD.
    """
    import os, io, datetime
    from flask import send_file, request
    from openpyxl import load_workbook

    # template filename desde query (fallback)
    template_name = request.args.get('template', 'Copia_INVENTARIO_PISO.xlsx')
    base_dir = os.path.dirname(__file__)
    tpl_dir = os.path.join(base_dir, 'excel_templates')
    tpl_path = os.path.join(tpl_dir, template_name)

    if not os.path.isfile(tpl_path):
        return f"Template not found: {tpl_path}", 404

    # obtener datos desde modelos o BD
    try:
        if 'find_transactions' in globals() and callable(find_transactions):
            tx_docs = find_transactions(aggregate_by_product=True, group_by_code=True)
            print("Loaded transactions via find_transactions(), count:", len(tx_docs))
        else:
            db = get_database()
            tx_docs = list(db.transactions.find().limit(1000))
    except Exception as e:
        print("Error loading transactions:", e)
        tx_docs = []

    # convertir lista de documentos en diccionario por código
    tx_map = {}
    for t in tx_docs:
        code = str(t.get('codigo') or t.get('code') or '').strip()
        if not code:
            continue

        try:
            total_val = str(t.get('total', ''))
            total_val = total_val.split('.')
            packd = t.get('packs', '')
            if len(total_val) > 1:
                total_val = total_val[0] + ',' + str(int(round(((int(total_val[1]) / (10 ** len(total_val[1]))) * packd), 1)))
            else:
                total_val = total_val[0]
        except Exception as e:
            print(f"Error processing total for {code}: {e}")
            total_val = 'Error'

        tx_map[code] = {
            'product': t.get('product', ''),
            'total_val': total_val
        }

    # cargar plantilla y rellenar
    wb = load_workbook(tpl_path)
    if 'Datos' not in wb.sheetnames:
        return "Sheet 'Datos' not found in template", 400

    ws = wb['Datos']

    # Asume encabezado en fila 6 y que los códigos empiezan en columna 1
    start_row = 6
    code_col = 1
    product_col = 2
    total_col = 4

    row = start_row
    while True:
        code_cell = ws.cell(row=row, column=code_col)
        code_value = str(code_cell.value).strip() if code_cell.value else ''
        if not code_value:  # si no hay más códigos, salimos
            break

        # buscar si existe ese código en la BD
        if code_value in tx_map:
            data = tx_map[code_value]
            ws.cell(row=row, column=product_col, value=data['product'])
            ws.cell(row=row, column=total_col, value=data['total_val'])
        else:
            # si no está en la BD, puedes limpiar la celda o dejarla como está
            ws.cell(row=row, column=total_col, value='0')

        row += 1

    # generar archivo en memoria y devolver
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename_out = f"report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename_out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/inventory/<id>')
def api_inventory(id):
    """Devuelve datos básicos del item de inventario por _id o por campo 'id' y suma de totales hoy."""
    try:
        db = None
        try:
            db = get_database()
        except Exception:
            pass
        if db is None and 'mongo' in globals() and mongo:
            db = mongo.db
        if db is None:
            return jsonify({'error': 'db_unavailable'}), 500

        # buscar por ObjectId primero, luego por campo 'id'
        doc = None
        try:
            oid = ObjectId(id)
            doc = db.inventory.find_one({'_id': oid}, {'product':1, 'code':1, 'shelves':1, 'floors':1, 'packs':1})
        except Exception:
            doc = db.inventory.find_one({'id': id}, {'product':1, 'code':1, 'shelves':1, 'floors':1, 'packs':1})

        if not doc:
            return jsonify({'error': 'not_found'}), 404

        # calcular rango del día según zona Colombia
        tz = ZoneInfo("America/Bogota")
        now_tz = datetime.datetime.now(tz)
        start = now_tz.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + datetime.timedelta(days=1)

        # pipeline: match por fecha y por inventory referencia o por product+codigo, luego sumar total
        match_cond = {
            '$and': [
                {'date': {'$gte': start, '$lt': end}},
                {'$or': [
                    {'inventory_id': doc.get('_id')},
                    {'$and': [{'product': doc.get('product')}, {'codigo': doc.get('code', '')}]}
                ]}
            ]
        }

        pipeline = [
            {'$match': match_cond},
            {'$group': {
                '_id': None,
                'total_sum': {'$sum': {'$ifNull': ['$total', 0]}},
                'count': {'$sum': 1}
            }}
        ]

        agg = list(db.transactions.aggregate(pipeline))
        total_sum = 0
        count = 0
        if agg:
            total_sum = agg[0].get('total_sum', 0) or 0
            count = agg[0].get('count', 0) or 0

        # serializar fields y ObjectId
        out = {}
        for k, v in doc.items():
            if isinstance(v, ObjectId):
                out[k] = str(v)
            else:
                out[k] = v

        try:
            total_sum = str(total_sum)
            total_val = total_sum.split('.')
            packd = doc.get('packs', '')
            if len(total_val) > 1:
                total_sum = total_val[0] + ',' + str( int(round(((int(total_val[1])/(10**len(total_val[1]))) * packd), 1)) )
            else:
                total_sum = total_val[0]
                total_sum = float(total_sum)

        except Exception as e:
            
            try:
                total_sum = float(total_sum)
            except Exception:
                total_sum = 0

        out.update({'transactions_today_sum': total_sum, 'transactions_today_count': int(count)})
        return jsonify(out)
    except Exception as e:
        return jsonify({'error': 'exception', 'msg': str(e)}), 500

if __name__ == '__main__':
    # comprobación al arrancar en consola
    ok, err = verify_mongo_connection()
    if ok:
        print("MongoDB connection OK")
    else:
        print("MongoDB connection FAILED:", err)
    app.run(debug=True)